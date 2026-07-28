import os
import time
import re
import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook
from geopy.geocoders import Nominatim

# ID Google Sheets & File Credentials Anda
SPREADSHEET_ID = "1z7WQdXzYZRejDFdNrZITFJITwlFoCLxXNIVSaWnQkos"  # Ganti jika ada ID Google Sheets baru
CREDENTIALS_FILE = "credentials.json"

# Inisialisasi Map Geocoder
geolocator = Nominatim(user_agent="odp_bth_web_generator_app_v2")

def get_google_sheet():
    """Koneksi ke Google Sheets (Mendukung Cloud Secrets & File Lokal)."""
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    if "gcp_service_account" in st.secrets:
        creds_dict = dict(st.secrets["gcp_service_account"])
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    else:
        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=scopes)
        
    client = gspread.authorize(creds)
    sheet = client.open_by_key(SPREADSHEET_ID).sheet1
    return sheet

def parse_koordinat(raw_text):
    """
    Validasi dan Parse Format Koordinat:
    - -6.294005, 107.330222
    - -6.230508° 107.360487°
    - -6.230508°107.360487°
    - -6.230508 107.360487
    - -6.230508,107.360487
    """
    if not raw_text:
        return None, None
    
    # Pisahkan simbol derajat dan koma menjadi spasi
    clean_text = str(raw_text).replace('°', ' ').replace(',', ' ').strip()
    parts = re.findall(r'[-+]?\d+\.\d+|[-+]?\d+', clean_text)
    
    if len(parts) == 2:
        try:
            lat = float(parts[0])
            lng = float(parts[1])
            # Memastikan nilai lat/lng dalam batas koordinat Bumi
            if -90 <= lat <= 90 and -180 <= lng <= 180:
                return lat, lng
        except ValueError:
            pass
    return None, None

def get_area_code(lat, lon):
    """Mencari nama kelurahan/wilayah berdasarkan koordinat & mengonversi ke 3 huruf kode area."""
    try:
        # Nominatim reverse lookup
        location = geolocator.reverse((lat, lon), timeout=15)
        if location:
            address = location.raw.get('address', {})
            
            # Cari nama wilayah dari level paling spesifik ke level yang lebih luas
            nama_wilayah = (
                address.get('village') or 
                address.get('suburb') or 
                address.get('neighbourhood') or 
                address.get('hamlet') or 
                address.get('subdistrict') or
                address.get('town') or
                address.get('city_district') or
                address.get('municipality') or
                address.get('county') or
                None
            )
            
            if nama_wilayah:
                # Bersihkan kata prefiks yang umum
                clean_name = (
                    nama_wilayah
                    .replace('Kelurahan', '')
                    .replace('Desa', '')
                    .replace('Kecamatan', '')
                    .replace('Kabupaten', '')
                    .strip()
                    .upper()
                )
                
                # Pemetaan Kode Area Spesifik
                if "SUKAMERTA" in clean_name:
                    return "SMT", nama_wilayah
                elif "BALONG" in clean_name or "BALONGSARI" in clean_name:
                    return "BLS", nama_wilayah
                elif "PASAWAHAN" in clean_name:
                    return "PSW", nama_wilayah
                elif "PALUMBONSARI" in clean_name:
                    return "PAL", nama_wilayah
                
                # Ambil 3 huruf pertama nama wilayah jika tidak ada di kondisi khusus
                kode_3_huruf = re.sub(r'[^A-Z]', '', clean_name)[:3]
                if len(kode_3_huruf) == 3:
                    return kode_3_huruf, nama_wilayah
                elif len(clean_name) >= 3:
                    return clean_name[:3], nama_wilayah
                    
    except Exception:
        pass
    
    # Jika benar-benar tidak terdeteksi oleh peta
    return "GEN", "Umum"

def load_master_database_gsheet(sheet):
    """Membaca Master Database langsung dari Google Sheets Online."""
    records = sheet.get_all_values()
    max_odp_counter = {}
    existing_labels = set()
    existing_locations_map = {}

    if len(records) == 0:
        sheet.append_row(["Kode ODP", "Nama Kelurahan", "Nomor ODC", "Koordinat", "Keterangan"])
        return max_odp_counter, existing_labels, existing_locations_map

    for row in records[1:]:
        if not row or not row[0]:
            continue
        
        kode_odp = str(row[0]).strip()
        odc = str(row[2]).strip() if len(row) > 2 else ""
        koordinat = str(row[3]).strip() if len(row) > 3 else ""

        existing_labels.add(kode_odp)
        if koordinat and odc:
            location_key = f"{koordinat}_{odc}"
            existing_locations_map[location_key] = kode_odp

        parts = kode_odp.split('_')
        if len(parts) >= 5:
            area = parts[2]
            odc_num = parts[3]
            try:
                odp = int(parts[4])
                key = f"{area}_{odc_num}"
                if key not in max_odp_counter or odp > max_odp_counter[key]:
                    max_odp_counter[key] = odp
            except ValueError:
                pass

    return max_odp_counter, existing_labels, existing_locations_map

def save_to_google_sheet(sheet, new_records):
    """Menambahkan data hasil generate baru ke baris paling bawah di Google Sheets."""
    rows_to_append = []
    for r in new_records:
        rows_to_append.append([
            r["kode_odp"],
            r["kelurahan"],
            r["odc"],
            r["koordinat"],
            r["keterangan"]
        ])
    if rows_to_append:
        sheet.append_rows(rows_to_append)

def generate_process(items_to_process):
    sheet = get_google_sheet()
    max_odp_counter, existing_labels, existing_locations_map = load_master_database_gsheet(sheet)
    results = []

    progress_bar = st.progress(0)
    total = len(items_to_process)

    for idx, item in enumerate(items_to_process):
        lat, lng = parse_koordinat(item["koordinat_raw"])
        
        # 1. KONDISI: KOORDINAT NOT VALID (Format Salah / Bukan Angka Koordinat Bumi)
        if lat is None or lng is None:
            results.append({
                "kode_odp": "",
                "kelurahan": "",
                "odc": "",
                "koordinat": "",
                "keterangan": "Koordinat  not valid"
            })
            progress_bar.progress((idx + 1) / total)
            continue

        odc_num = str(item["odc_raw"]).strip().zfill(3)
        formatted_koordinat = f"{lat}, {lng}"
        location_key = f"{formatted_koordinat}_{odc_num}"

        # 2. KONDISI: KOORDINAT DUPLIKAT (Sudah ada di Google Sheets)
        if location_key in existing_locations_map:
            results.append({
                "kode_odp": "",
                "kelurahan": "",
                "odc": "",
                "koordinat": formatted_koordinat,
                "keterangan": "Koordinat duplikat"
            })
            time.sleep(0.3)
            progress_bar.progress((idx + 1) / total)
            continue

        # 3. KONDISI: KOORDINAT VALID (Proses Pembuatan Kode ODP Baru)
        kode_area, nama_kelurahan = get_area_code(lat, lng)

        key = f"{kode_area}_{odc_num}"
        while True:
            last_odp_num = max_odp_counter.get(key, 0)
            next_odp_num = last_odp_num + 1
            max_odp_counter[key] = next_odp_num
            
            odp_str = str(next_odp_num).zfill(3)
            kode_odp_final = f"ODP_BTH_{kode_area}_{odc_num}_{odp_str}"

            if kode_odp_final not in existing_labels:
                existing_labels.add(kode_odp_final)
                existing_locations_map[location_key] = kode_odp_final
                break

        record = {
            "kode_odp": kode_odp_final,
            "kelurahan": nama_kelurahan,
            "odc": odc_num,
            "koordinat": formatted_koordinat,
            "keterangan": "DONE"
        }

        results.append(record)
        time.sleep(1.0) # Delay API Geocoder
        progress_bar.progress((idx + 1) / total)

    save_to_google_sheet(sheet, results)
    return results

# ================= UI STREAMLIT =================
st.set_page_config(page_title="Generator ODP (Google Sheets)", layout="centered")

st.title("⚡ Generator Kode ODP BTH")
st.caption("Terhubung langsung dengan Master Database Google Sheets Cloud")

tab1, tab2 = st.tabs(["📝 Input Manual Single", "📁 Drop File Excel"])

items_to_process = []

with tab1:
    st.subheader("Opsi 1: Isian Manual")
    input_koordinat = st.text_input("Koordinat :", placeholder="Contoh: -6.293009, 107.330705")
    input_odc = st.text_input("ODC :", placeholder="Contoh: 001 atau 1")

with tab2:
    st.subheader("Opsi 2: Drop File Excel")
    uploaded_file = st.file_uploader("Drop file Excel input", type=["xlsx", "xls"])

st.markdown("---")

if st.button("🚀 GENERATE KODE ODP", type="primary"):
    if uploaded_file is not None:
        wb = load_workbook(uploaded_file)
        sheet = wb.active
        for row_idx in range(2, sheet.max_row + 1):
            koordinat_val = sheet.cell(row=row_idx, column=1).value
            odc_val = sheet.cell(row=row_idx, column=2).value
            if koordinat_val or odc_val is not None:
                items_to_process.append({"koordinat_raw": koordinat_val, "odc_raw": odc_val if odc_val is not None else ""})
        st.info(f"Membaca {len(items_to_process)} data dari file Excel...")
    
    elif input_koordinat or input_odc:
        items_to_process.append({"koordinat_raw": input_koordinat, "odc_raw": input_odc if input_odc else ""})
        st.info("Memproses 1 data dari isian manual...")
    
    else:
        st.warning("⚠️ Harap isi form Manual atau Unggah File Excel terlebih dahulu!")

    if items_to_process:
        with st.spinner("Sedang menghubungkan ke Google Sheets & memproses area..."):
            try:
                res = generate_process(items_to_process)
                st.success("✅ Selesai! Data otomatis tersimpan di Google Sheets Online.")
                st.subheader("Hasil Generate:")
                
                # Tampilan Tabel Hasil
                html_table = "<table border='1' style='border-collapse: collapse; width: 100%; text-align: left;'>"
                html_table += "<tr style='background-color: #f2f2f2;'><th>No</th><th>Kode ODP</th><th>Nama Kelurahan</th><th>Nomor ODC</th><th>Koordinat</th><th>Keterangan</th></tr>"
                for idx_row, row in enumerate(res, 1):
                    html_table += f"<tr><td>{idx_row}</td><td><b>{row['kode_odp']}</b></td><td>{row['kelurahan']}</td><td>{row['odc']}</td><td>{row['koordinat']}</td><td>{row['keterangan']}</td></tr>"
                html_table += "</table>"
                
                st.markdown(html_table, unsafe_allow_html=True)
                
            except Exception as e:
                st.error(f"Terjadi kesalahan saat menghubungkan ke Google Sheets: {e}")
